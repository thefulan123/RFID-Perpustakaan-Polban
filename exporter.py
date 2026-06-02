import os
import openpyxl
from datetime import datetime

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_pengunjung")
BULAN_INDONESIA = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]
HARI_INDONESIA = [
    "Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu",
]
HEADERS = ["No", "UID", "NIM", "Nama", "Tanggal", "Jam"]


def _get_paths(tanggal=None):
    if tanggal is None:
        tanggal = datetime.now()
    tahun = str(tanggal.year)
    bulan = f"{tanggal.month:02d}-{BULAN_INDONESIA[tanggal.month]}"
    hari = HARI_INDONESIA[tanggal.weekday()]
    sheet_name = f"{tanggal.day:02d}-{hari}"
    folder = os.path.join(BASE_DIR, tahun, bulan)
    filepath = os.path.join(folder, f"{bulan}.xlsx")
    return folder, filepath, sheet_name


def _ensure_folder(folder):
    if not os.path.exists(folder):
        os.makedirs(folder)


def _get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(HEADERS)
    return ws


def _get_next_row(ws):
    return ws.max_row + 1


def simpan_kunjungan(uid, nim, nama):
    now = datetime.now()
    folder, filepath, sheet_name = _get_paths(now)

    _ensure_folder(folder)

    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
    else:
        wb = openpyxl.Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

    ws = _get_or_create_sheet(wb, sheet_name)
    row_num = _get_next_row(ws)
    tanggal = now.strftime("%Y-%m-%d")
    jam = now.strftime("%H:%M:%S")
    ws.append([row_num - 1, uid, nim, nama, tanggal, jam])

    wb.save(filepath)
    return filepath
